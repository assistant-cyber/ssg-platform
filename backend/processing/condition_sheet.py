"""
Stained Glass Condition Sheet Generator
========================================
Parses photo shorthand descriptions and produces an openpyxl condition
spreadsheet.  Ported from populate_condition_sheet.py + build_templates.py.

No dependency on CompanyCam client or app/ package.

Three parsing modes:
  shorthand (default) — compact tokens: 1A w2 l1 b0 rot p 61pc 30x36
  ai                  — Claude Haiku (requires ANTHROPIC_API_KEY env var)
  hybrid              — shorthand first, AI fallback

Rubric: 0-1 = Good, 2 = Fair, 3-5 = Poor
"""
import json
import math
import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

# ─── PanelData ────────────────────────────────────────────────────────────────

@dataclass
class PanelData:
    window_num: str
    panel_letter: str = ""
    elevation: str = ""
    warping: Optional[int] = None
    lead_det: Optional[int] = None
    breaks: Optional[int] = None
    wood_rot: Optional[bool] = None
    paint_fail: Optional[bool] = None
    pieces: Optional[int] = None
    panel_w: Optional[float] = None
    panel_h: Optional[float] = None
    overall_w: Optional[float] = None
    overall_h: Optional[float] = None
    notes: str = ""
    is_overall_only: bool = False
    photo_url: str = ""      # path or URL (for optional vision piece counting)
    photo_url_web: str = ""


# ─── Shorthand regex patterns ─────────────────────────────────────────────────

_RE_WINDOW_ID    = re.compile(r'^(\d+)([a-zA-Z])?')
_RE_WARPING      = re.compile(r'\bw(\d)\b', re.I)
_RE_LEAD         = re.compile(r'\bl(\d)\b', re.I)
_RE_BREAKS       = re.compile(r'\bb(\d+)\b', re.I)
_RE_ROT          = re.compile(r'\brot\b', re.I)
_RE_PAINT        = re.compile(r'\bp\b(?!\w)', re.I)
_RE_PIECES       = re.compile(r'\b(\d+)\s*pc\b', re.I)
_RE_OVERALL_DIMS = re.compile(r'\bov\s*(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\b', re.I)
_RE_DIMS         = re.compile(r'\b(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)\b', re.I)

_ELEVATIONS = ["north", "south", "east", "west", "ne", "nw", "se", "sw"]


def _detect_shorthand(text: str) -> bool:
    """Return True if text contains at least one severity shorthand token."""
    return bool(_RE_WARPING.search(text) or _RE_LEAD.search(text))


def _extract_elevation(text: str) -> str:
    lower = text.lower()
    for elev in _ELEVATIONS:
        if re.search(r'\b' + elev + r'\b', lower):
            return elev.upper()
    return ""


# ─── SHORTHAND PARSER ─────────────────────────────────────────────────────────

def parse_shorthand(description: str) -> Optional[PanelData]:
    """Parse compact shorthand description into PanelData."""
    text = description.strip()
    if not text:
        return None

    m = _RE_WINDOW_ID.match(text)
    if not m:
        return None

    win_num = m.group(1)
    panel_letter = (m.group(2) or "").upper()
    rest = text[m.end():].strip()
    rest_clean = re.sub(r'^[\.\-\,\:\;]+\s*', '', rest)

    pd = PanelData(window_num=win_num, panel_letter=panel_letter)
    pd.elevation = _extract_elevation(rest_clean)

    # Overall dimensions (ov prefix)
    ov_m = _RE_OVERALL_DIMS.search(rest_clean)
    if ov_m:
        pd.overall_w = float(ov_m.group(1))
        pd.overall_h = float(ov_m.group(2))

    if not panel_letter:
        # No panel letter → this is a whole-window (overall-only) photo
        dims_m = _RE_DIMS.search(rest_clean)
        if dims_m and not ov_m:
            pd.overall_w = float(dims_m.group(1))
            pd.overall_h = float(dims_m.group(2))
        pd.is_overall_only = True
        return pd

    # Severity tokens
    wm = _RE_WARPING.search(rest_clean)
    if wm:
        pd.warping = int(wm.group(1))
    lm = _RE_LEAD.search(rest_clean)
    if lm:
        pd.lead_det = int(lm.group(1))
    bm = _RE_BREAKS.search(rest_clean)
    if bm:
        pd.breaks = int(bm.group(1))

    pd.wood_rot = bool(_RE_ROT.search(rest_clean))
    rest_no_pc = _RE_PIECES.sub('', rest_clean)
    pd.paint_fail = bool(_RE_PAINT.search(rest_no_pc))

    pcm = _RE_PIECES.search(rest_clean)
    if pcm:
        pd.pieces = int(pcm.group(1))

    # Panel dims
    if not ov_m:
        dims_m = _RE_DIMS.search(rest_clean)
        if dims_m:
            pd.panel_w = float(dims_m.group(1))
            pd.panel_h = float(dims_m.group(2))
    else:
        rest_no_ov = rest_clean[:ov_m.start()] + rest_clean[ov_m.end():]
        dims_m = _RE_DIMS.search(rest_no_ov)
        if dims_m:
            pd.panel_w = float(dims_m.group(1))
            pd.panel_h = float(dims_m.group(2))

    # Leftover text → notes
    leftover = rest_clean
    for pattern in [_RE_WARPING, _RE_LEAD, _RE_BREAKS, _RE_ROT,
                    _RE_PIECES, _RE_OVERALL_DIMS, _RE_DIMS]:
        leftover = pattern.sub('', leftover)
    leftover = _RE_PAINT.sub('', leftover)
    leftover = re.sub(r'\s+', ' ', leftover).strip(' ,.-;:')
    for word in _ELEVATIONS:
        leftover = re.sub(r'\b' + re.escape(word) + r'\b', '', leftover, flags=re.I)
    leftover = re.sub(r'\s+', ' ', leftover).strip(' ,.-;:')
    pd.notes = leftover

    return pd


# ─── AI PARSER ────────────────────────────────────────────────────────────────

_AI_SYSTEM_PROMPT = """You are a stained glass window condition data extractor.
Given a photo description from a field assessment, extract structured condition data.

Return ONLY valid JSON with these fields (use null for missing/unknown values):
{
  "elevation": "direction if mentioned: North, South, East, West, NE, NW, SE, SW",
  "warping": 0-5 integer severity (0=none, 1=very minor, 2=moderate, 3=significant, 4=severe, 5=critical),
  "lead_det": 0-5 integer severity for lead deterioration/sagging/cracking,
  "breaks": integer count of broken glass pieces (0 if none mentioned),
  "wood_rot": true/false whether wood rot or frame damage is mentioned,
  "paint_fail": true/false whether failing paint, caulking, or peeling is mentioned,
  "pieces": integer count of glass pieces in the panel (null if not mentioned),
  "panel_w": width in inches (null if not mentioned),
  "panel_h": height in inches (null if not mentioned),
  "overall_w": overall window width in inches (null if not mentioned),
  "overall_h": overall window height in inches (null if not mentioned)
}

Severity scale: 0=none, 1=very minor, 2=moderate, 3=significant, 4=severe, 5=critical.
Return ONLY the JSON object, nothing else."""


def _get_anthropic_client():
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY not set. Add it to your .env file to use AI parsing."
        )
    try:
        import anthropic
    except ImportError:
        raise ImportError("Install the 'anthropic' package: pip install anthropic")
    return anthropic.Anthropic(api_key=api_key)


def parse_ai(description: str) -> Optional[PanelData]:
    """Parse a freeform description using Claude Haiku."""
    text = description.strip()
    if not text:
        return None

    m = _RE_WINDOW_ID.match(text)
    if not m:
        return None

    win_num = m.group(1)
    panel_letter = (m.group(2) or "").upper()
    rest = text[m.end():].strip()
    rest = re.sub(r'^[\.\-\,\:\;]+\s*', '', rest)

    if len(rest) < 3:
        pd = PanelData(window_num=win_num, panel_letter=panel_letter)
        dims_m = _RE_DIMS.search(rest)
        if dims_m:
            if not panel_letter:
                pd.overall_w = float(dims_m.group(1))
                pd.overall_h = float(dims_m.group(2))
                pd.is_overall_only = True
            else:
                pd.panel_w = float(dims_m.group(1))
                pd.panel_h = float(dims_m.group(2))
        return pd

    client = _get_anthropic_client()
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        system=_AI_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": rest}],
    )

    try:
        result_text = response.content[0].text.strip()
        if result_text.startswith("```"):
            result_text = re.sub(r'^```\w*\n?', '', result_text)
            result_text = re.sub(r'\n?```$', '', result_text)
        data = json.loads(result_text)
    except (json.JSONDecodeError, IndexError) as e:
        print(f"  Warning: AI returned unparseable response for '{text[:60]}': {e}")
        return PanelData(window_num=win_num, panel_letter=panel_letter, notes=rest)

    pd = PanelData(
        window_num=win_num,
        panel_letter=panel_letter,
        elevation=data.get("elevation") or "",
        warping=data.get("warping"),
        lead_det=data.get("lead_det"),
        breaks=data.get("breaks"),
        wood_rot=data.get("wood_rot"),
        paint_fail=data.get("paint_fail"),
        pieces=data.get("pieces"),
        panel_w=data.get("panel_w"),
        panel_h=data.get("panel_h"),
        overall_w=data.get("overall_w"),
        overall_h=data.get("overall_h"),
    )

    if not panel_letter and pd.warping is None and pd.lead_det is None:
        pd.is_overall_only = True

    return pd


# ─── HYBRID PARSER ────────────────────────────────────────────────────────────

def parse_hybrid(description: str) -> Optional[PanelData]:
    """Try shorthand first; fall back to AI if no severity tokens found."""
    text = description.strip()
    if not text:
        return None

    m = _RE_WINDOW_ID.match(text)
    if not m:
        return None

    rest = text[m.end():].strip()
    rest = re.sub(r'^[\.\-\,\:\;]+\s*', '', rest)

    if _detect_shorthand(rest):
        return parse_shorthand(text)
    else:
        return parse_ai(text)


def select_parser(mode: str):
    if mode == "shorthand":
        return parse_shorthand
    elif mode == "ai":
        return parse_ai
    elif mode == "hybrid":
        return parse_hybrid
    else:
        raise ValueError(f"Unknown parsing mode: {mode}")


# ─── OpenCV piece counter ─────────────────────────────────────────────────────

def _count_pieces_opencv(image_bytes: bytes) -> Optional[Dict]:
    """Count glass pieces using OpenCV contour detection.

    Returns dict with 'piece_count' and 'confidence', or None on failure.
    Guarded by try/except ImportError so OpenCV is optional.
    """
    try:
        import cv2
        import numpy as np
    except ImportError:
        return None

    img_array = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    if img is None:
        return None

    h, w = img.shape[:2]
    total_pixels = h * w

    # Resize to ~1MP for consistent processing
    target_pixels = 1_000_000
    if total_pixels > target_pixels * 1.5:
        scale = (target_pixels / total_pixels) ** 0.5
        img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        h, w = img.shape[:2]
        total_pixels = h * w

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)

    thresh = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 19, 6
    )

    kernel_close = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel_close, iterations=1)
    kernel_open = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    cleaned = cv2.morphologyEx(closed, cv2.MORPH_OPEN, kernel_open, iterations=1)

    contours, _ = cv2.findContours(cleaned, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    min_area = total_pixels * 0.001
    max_area = total_pixels * 0.30
    valid_areas = [cv2.contourArea(c) for c in contours
                   if min_area < cv2.contourArea(c) < max_area]

    piece_count = len(valid_areas)

    if piece_count < 5 or piece_count > 200:
        confidence = "low"
    else:
        arr = np.array(valid_areas)
        cv_ratio = np.std(arr) / np.mean(arr) if arr.mean() > 0 else 99
        confidence = "high" if cv_ratio < 1.5 else "medium"

    return {"piece_count": piece_count, "confidence": confidence}


# ─── Excel builder (inlined from build_templates + populate_condition_sheet) ──

def _build_excel(
    panels: List[PanelData],
    output_path: str,
    project_name: str = "",
    flavor: str = "stained",
) -> str:
    """Build a complete condition spreadsheet from parsed panel data.

    Creates the workbook structure (from build_templates.py) and populates it
    with panel data (from populate_condition_sheet.py) in one pass.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.comments import Comment

    # ── Style constants ─────────────────────────────────────────────────────
    NAVY        = "1F3864"
    NAVY_LIGHT  = "2E5A9A"
    GOLD        = "C9A227"
    CREAM       = "FFF8E7"
    BAND_ALT    = "F5F7FB"
    WHITE       = "FFFFFF"
    TEXT_DARK   = "1A1A1A"
    GOOD_GREEN  = "C6EFCE"
    FAIR_YELLOW = "FFEB9C"
    POOR_RED    = "F8B4B4"

    THIN = Side(style="thin", color="B8BFCC")
    BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    FONT_NAME = "Arial"
    DATA_START = 5

    # ── Merge / group panel data ────────────────────────────────────────────
    windows: Dict[str, Dict] = {}
    for pd in panels:
        wn = pd.window_num
        if wn not in windows:
            windows[wn] = {"overall": None, "panels": []}
        if pd.is_overall_only:
            existing = windows[wn]["overall"]
            if existing is None:
                windows[wn]["overall"] = pd
            else:
                if pd.overall_w and not existing.overall_w:
                    existing.overall_w = pd.overall_w
                if pd.overall_h and not existing.overall_h:
                    existing.overall_h = pd.overall_h
                if pd.elevation and not existing.elevation:
                    existing.elevation = pd.elevation
        else:
            existing_panel = next(
                (p for p in windows[wn]["panels"] if p.panel_letter == pd.panel_letter),
                None,
            )
            if existing_panel and pd.panel_letter:
                # Merge: keep whichever field has data
                if pd.warping is not None and existing_panel.warping is None:
                    existing_panel.warping = pd.warping
                if pd.lead_det is not None and existing_panel.lead_det is None:
                    existing_panel.lead_det = pd.lead_det
                if pd.breaks is not None and existing_panel.breaks is None:
                    existing_panel.breaks = pd.breaks
                if pd.wood_rot is not None and existing_panel.wood_rot is None:
                    existing_panel.wood_rot = pd.wood_rot
                if pd.paint_fail is not None and existing_panel.paint_fail is None:
                    existing_panel.paint_fail = pd.paint_fail
                if pd.pieces is not None and existing_panel.pieces is None:
                    existing_panel.pieces = pd.pieces
                if pd.panel_w and not existing_panel.panel_w:
                    existing_panel.panel_w = pd.panel_w
                if pd.panel_h and not existing_panel.panel_h:
                    existing_panel.panel_h = pd.panel_h
            else:
                windows[wn]["panels"].append(pd)

    sorted_windows = sorted(windows.keys(), key=lambda x: int(x) if x.isdigit() else 999)
    for wn in sorted_windows:
        windows[wn]["panels"].sort(key=lambda p: p.panel_letter or "Z")

    # Build window_panel_map
    window_panel_map: Dict[str, List[str]] = {}
    for wn in sorted_windows:
        panel_list = windows[wn]["panels"]
        if panel_list:
            window_panel_map[wn] = [pd.panel_letter or chr(ord("A") + i)
                                    for i, pd in enumerate(panel_list)]
        else:
            window_panel_map[wn] = ["A"]

    # ── Create workbook ─────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Window Conditions"

    if flavor == "stained":
        title = "Stained Glass Window Condition Assessment"
        metric_cols = [
            ("Warping", "0-5 severity scale"),
            ("Lead Deterioration", "0-5 severity scale"),
            ("Glass Breaks", "Count of broken pieces"),
            ("Wood Rot / Frame Damage", "Yes / No"),
            ("Failing Paint / Caulk", "Yes / No"),
        ]
    else:
        title = "Dalle de Verre Window Condition Assessment"
        metric_cols = [
            ("Cracking Epoxy", "0-5 severity scale"),
            ("Glass / Epoxy Separation", "0-5 severity scale"),
            ("Glass Breaks", "Count of broken pieces"),
            ("Water Entry", "Yes / No"),
            ("Failing Caulk", "Yes / No"),
        ]

    col_headers = [
        "Window /\nPanel", "Elevation",
        metric_cols[0][0], metric_cols[1][0], metric_cols[2][0],
        metric_cols[3][0], metric_cols[4][0],
        "# Pieces", "Panel W\n(in)", "Panel H\n(in)", "Panel\nSqFt",
        "Good\nSqFt", "Fair\nSqFt", "Poor\nSqFt",
        "Overall W\n(in)", "Overall H\n(in)", "Overall\nSqFt",
        "Notes", "_Cond", "_IsWindow",
    ]
    col_widths = [11, 10, 12, 14, 10, 14, 14, 9, 10, 10, 10, 10, 10, 10, 11, 11, 10, 28, 2, 2]

    # ── Row 1: Title banner ─────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name=FONT_NAME, bold=True, size=18, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=NAVY)
    ws.row_dimensions[1].height = 34

    # ── Row 2: Sub-banner ───────────────────────────────────────────────────
    ws.merge_cells("A2:C2"); ws["A2"].value = "Project / Church:"
    ws.merge_cells("D2:G2"); ws["D2"].value = project_name or ""
    ws.merge_cells("H2:J2"); ws["H2"].value = "Assessment Date:"
    ws.merge_cells("K2:M2"); ws["K2"].value = ""
    ws.merge_cells("N2:O2"); ws["N2"].value = "Assessor:"
    ws.merge_cells("P2:R2"); ws["P2"].value = ""
    for addr in ["A2", "H2", "N2"]:
        ws[addr].font = Font(name=FONT_NAME, bold=True, size=11, color=WHITE)
        ws[addr].fill = PatternFill("solid", start_color=NAVY_LIGHT)
        ws[addr].alignment = Alignment(horizontal="right", vertical="center")
    for addr in ["D2", "K2", "P2"]:
        ws[addr].font = Font(name=FONT_NAME, size=11, color=TEXT_DARK)
        ws[addr].fill = PatternFill("solid", start_color=CREAM)
        ws[addr].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Row 3: Group headers ────────────────────────────────────────────────
    groups = [
        ("A3:B3", "IDENTIFICATION"),
        ("C3:E3", "PANEL CONDITION"),
        ("F3:G3", "EXTERIOR / FRAME"),
        ("H3:K3", "PANEL DETAILS"),
        ("L3:N3", "SQFT BY CONDITION"),
        ("O3:Q3", "OVERALL WINDOW"),
        ("R3:R3", ""),
    ]
    for rng, label in groups:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", start_color=NAVY_LIGHT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # ── Row 4: Column headers ───────────────────────────────────────────────
    for i, (header, width) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=4, column=i, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", start_color=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 38

    # Comments on metric columns
    for col_idx, (_, help_text) in zip([3, 4, 5], metric_cols[:3]):
        ws.cell(row=4, column=col_idx).comment = Comment(help_text, "SSG")

    # ── Data rows ────────────────────────────────────────────────────────────
    row = DATA_START
    window_row_map = []

    for win_idx, wn in enumerate(sorted_windows):
        win_data = windows[wn]
        panel_letters = window_panel_map[wn]
        num_panels = len(panel_letters)
        summary_row = row

        # --- Summary row ---
        ws.cell(row=row, column=1, value=str(wn))

        all_panels = win_data["panels"]
        overall = win_data["overall"]

        # Elevation from first panel or overall
        if all_panels:
            elev = all_panels[0].elevation
        elif overall:
            elev = overall.elevation
        else:
            elev = ""
        if elev:
            ws.cell(row=summary_row, column=2, value=elev)

        # Overall dims
        ow = oh = None
        if overall:
            ow, oh = overall.overall_w, overall.overall_h
        if not ow and all_panels:
            for pp in all_panels:
                if pp.overall_w and pp.overall_h:
                    ow, oh = pp.overall_w, pp.overall_h
                    break
        if ow:
            ws.cell(row=summary_row, column=15, value=ow)
        if oh:
            ws.cell(row=summary_row, column=16, value=oh)
        if ow and oh:
            ws.cell(row=summary_row, column=17, value=math.ceil(ow * oh / 144))

        # Window-level rot/paint (aggregate from panels)
        has_rot = any(p.wood_rot for p in all_panels if p.wood_rot is not None)
        has_paint = any(p.paint_fail for p in all_panels if p.paint_fail is not None)
        if overall:
            if overall.wood_rot:
                has_rot = True
            if overall.paint_fail:
                has_paint = True
        if has_rot:
            ws.cell(row=summary_row, column=6, value="Yes")
        if has_paint:
            ws.cell(row=summary_row, column=7, value="Yes")

        # Window condition formula (worst of panels below)
        pf = row + 1
        pl = row + num_panels
        cond_rng = f"S{pf}:S{pl}"
        ws.cell(row=row, column=19).value = (
            f'=IF(COUNTIF({cond_rng},"Poor")>0,"Poor",'
            f'IF(COUNTIF({cond_rng},"Fair")>0,"Fair",'
            f'IF(COUNTIF({cond_rng},"Good")>0,"Good","")))'
        )
        ws.cell(row=row, column=20, value=1)  # _IsWindow = 1

        # Style summary row
        for col in range(1, 21):
            c = ws.cell(row=summary_row, column=col)
            c.font = Font(name=FONT_NAME, bold=True, size=11, color=WHITE)
            c.fill = PatternFill("solid", start_color=NAVY_LIGHT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER_ALL
        ws.row_dimensions[summary_row].height = 22

        panel_rows_for_validation = []

        # --- Panel rows ---
        for p_idx, pd in enumerate(all_panels[:num_panels]):
            row += 1
            panel_rows_for_validation.append(row)
            label = f"{wn}{pd.panel_letter}" if pd.panel_letter else f"{wn}"
            ws.cell(row=row, column=1, value=label)
            if pd.elevation:
                ws.cell(row=row, column=2, value=pd.elevation)
            elif elev:
                ws.cell(row=row, column=2, value=elev)

            if pd.warping is not None:
                ws.cell(row=row, column=3, value=pd.warping)
            if pd.lead_det is not None:
                ws.cell(row=row, column=4, value=pd.lead_det)
            if pd.breaks is not None:
                ws.cell(row=row, column=5, value=pd.breaks)
            if pd.pieces is not None:
                ws.cell(row=row, column=8, value=pd.pieces)
            if pd.panel_w is not None:
                ws.cell(row=row, column=9, value=pd.panel_w)
            if pd.panel_h is not None:
                ws.cell(row=row, column=10, value=pd.panel_h)
            if pd.panel_w is not None and pd.panel_h is not None:
                ws.cell(row=row, column=11, value=math.ceil(pd.panel_w * pd.panel_h / 144))

            # Hidden condition formula (S = col 19)
            s1, s2 = f"C{row}", f"D{row}"
            ws.cell(row=row, column=19, value=(
                f'=IFERROR(IF(AND(NOT(ISNUMBER({s1})),NOT(ISNUMBER({s2}))),"",'
                f'IF(MAX(IFERROR({s1},0),IFERROR({s2},0))>=3,"Poor",'
                f'IF(MAX(IFERROR({s1},0),IFERROR({s2},0))=2,"Fair","Good"))),"")'
            ))

            # Good/Fair/Poor SqFt formulas
            ws.cell(row=row, column=12, value=f'=IF(S{row}="Good",K{row},"")')
            ws.cell(row=row, column=12).number_format = "0"
            ws.cell(row=row, column=13, value=f'=IF(S{row}="Fair",K{row},"")')
            ws.cell(row=row, column=13).number_format = "0"
            ws.cell(row=row, column=14, value=f'=IF(S{row}="Poor",K{row},"")')
            ws.cell(row=row, column=14).number_format = "0"

            ws.cell(row=row, column=20, value=0)  # _IsWindow = 0

            if pd.notes:
                ws.cell(row=row, column=18, value=pd.notes)

            # Style panel row
            band = CREAM if (win_idx % 2 == 0) else BAND_ALT
            for col in range(1, 21):
                c = ws.cell(row=row, column=col)
                c.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
                c.fill = PatternFill("solid", start_color=band)
                c.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=(col == 18)
                )
                c.border = BORDER_ALL
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True, size=10, color=NAVY)
            ws.cell(row=row, column=18).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.row_dimensions[row].height = 20

        window_row_map.append((summary_row, panel_rows_for_validation))
        row += 1

    LAST_DATA_ROW = row - 1

    # ── Data validation ──────────────────────────────────────────────────────
    sev_dv = DataValidation(type="list", formula1='"0,1,2,3,4,5"',
                            allow_blank=True, showDropDown=False)
    sev_dv.errorTitle = "Invalid severity"
    sev_dv.error = "Enter 0 (none) to 5 (severe)."
    yn_dv = DataValidation(type="list", formula1='"Yes,No"',
                           allow_blank=True, showDropDown=False)
    elev_dv = DataValidation(
        type="list", formula1='"North,South,East,West,NE,NW,SE,SW"',
        allow_blank=True, showDropDown=False,
    )
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(yn_dv)
    ws.add_data_validation(elev_dv)

    for _summary, panel_rows_v in window_row_map:
        yn_dv.add(f"F{_summary}"); yn_dv.add(f"G{_summary}")
        for pr in panel_rows_v:
            sev_dv.add(f"C{pr}"); sev_dv.add(f"D{pr}")
            elev_dv.add(f"B{pr}")

    # ── Conditional formatting ───────────────────────────────────────────────
    for col_letter, color in [("L", GOOD_GREEN), ("M", FAIR_YELLOW), ("N", POOR_RED)]:
        rng = f"{col_letter}{DATA_START}:{col_letter}{LAST_DATA_ROW}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThan", formula=["0"],
                fill=PatternFill("solid", start_color=color),
            ),
        )

    cond_range = f"S{DATA_START}:S{LAST_DATA_ROW}"
    for val, color in [("Good", GOOD_GREEN), ("Fair", FAIR_YELLOW), ("Poor", POOR_RED)]:
        ws.conditional_formatting.add(
            cond_range,
            CellIsRule(
                operator="equal", formula=[f'"{val}"'],
                fill=PatternFill("solid", start_color=color),
            ),
        )

    # Hide helper columns S, T
    ws.column_dimensions["S"].hidden = True
    ws.column_dimensions["T"].hidden = True
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.print_title_rows = "1:4"

    # ── Summary / Dashboard sheet ────────────────────────────────────────────
    ds = wb.create_sheet("Summary")
    ds.sheet_view.showGridLines = False

    ds.merge_cells("A1:E1")
    ds["A1"].value = title.replace("Condition Assessment", "") + "-- Summary"
    ds["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=WHITE)
    ds["A1"].fill = PatternFill("solid", start_color=NAVY)
    ds["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ds.row_dimensions[1].height = 30

    cs = "'Window Conditions'"
    cond_rng_s   = f"{cs}!S{DATA_START}:S{LAST_DATA_ROW}"
    pieces_rng   = f"{cs}!H{DATA_START}:H{LAST_DATA_ROW}"
    breaks_rng   = f"{cs}!E{DATA_START}:E{LAST_DATA_ROW}"
    ov_sqft      = f"{cs}!Q{DATA_START}:Q{LAST_DATA_ROW}"
    flag_rng     = f"{cs}!T{DATA_START}:T{LAST_DATA_ROW}"
    good_sqft    = f"{cs}!L{DATA_START}:L{LAST_DATA_ROW}"
    fair_sqft    = f"{cs}!M{DATA_START}:M{LAST_DATA_ROW}"
    poor_sqft    = f"{cs}!N{DATA_START}:N{LAST_DATA_ROW}"
    rot_rng      = f"{cs}!F{DATA_START}:F{LAST_DATA_ROW}"
    paint_rng    = f"{cs}!G{DATA_START}:G{LAST_DATA_ROW}"

    overview_items = [
        ("Total Windows",       f'=SUMPRODUCT(({flag_rng}=1)*({cond_rng_s}<>""))'),
        ("Total Panels Logged", f'=SUMPRODUCT(({flag_rng}=0)*({cond_rng_s}<>""))'),
        ("Total # Pieces",      f'=SUMIFS({pieces_rng},{flag_rng},0)'),
        ("Total Overall SqFt",  f'=SUMIFS({ov_sqft},{flag_rng},1)'),
    ]

    r = 3
    for label, formula in overview_items:
        lc = ds.cell(row=r, column=1, value=label)
        lc.font = Font(name=FONT_NAME, bold=True, size=11, color=WHITE)
        lc.fill = PatternFill("solid", start_color=NAVY_LIGHT)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER_ALL
        vc = ds.cell(row=r, column=2, value=formula)
        vc.font = Font(name=FONT_NAME, bold=True, size=11, color=TEXT_DARK)
        vc.fill = PatternFill("solid", start_color=CREAM)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = BORDER_ALL
        vc.number_format = "#,##0"
        ds.row_dimensions[r].height = 22
        r += 1

    ds.column_dimensions["A"].width = 28
    ds.column_dimensions["B"].width = 20
    for col in "CDE":
        ds.column_dimensions[col].width = 14

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


# ─── Main DB entry point ──────────────────────────────────────────────────────

def generate_condition_sheet_from_db(
    project_id: str,
    project_name: str,
    photos: List[Dict],
    output_path: str,
    mode: str = "shorthand",
    count_pieces: bool = False,
    flavor: str = "stained",
) -> str:
    """Parse photo notes and generate a condition spreadsheet.

    Args:
        project_id:   SSG project UUID (unused internally, kept for API consistency).
        project_name: Church / project name for the spreadsheet header.
        photos:       List of dicts with keys: id, notes, local_path,
                      window_number, panel_letter, elevation, …
        output_path:  Where to write the .xlsx file.
        mode:         "shorthand" | "ai" | "hybrid"
        count_pieces: Whether to auto-count glass pieces via OpenCV.
        flavor:       "stained" | "dalle"

    Returns:
        The output_path on success.
    """
    parser_fn = select_parser(mode)
    panels: List[PanelData] = []
    skipped = 0

    for photo in photos:
        notes = (photo.get("notes") or "").strip()

        # If notes doesn't start with a window number, try constructing from DB fields
        if not notes or not _RE_WINDOW_ID.match(notes):
            win = photo.get("window_number", "")
            pan = photo.get("panel_letter", "")
            if win:
                synthetic = f"{win}{pan or ''} {notes}".strip()
                notes = synthetic
            else:
                skipped += 1
                continue

        if not _RE_WINDOW_ID.match(notes):
            skipped += 1
            continue

        try:
            pd = parser_fn(notes)
            if pd:
                # Override elevation from DB if not in parsed text
                if not pd.elevation and photo.get("elevation"):
                    pd.elevation = photo["elevation"]
                # Set photo path for optional piece counting
                pd.photo_url = photo.get("local_path", "")
                panels.append(pd)
        except Exception as e:
            print(f"  Warning: Failed to parse photo {photo.get('id', '?')}: {e}")
            skipped += 1

    # Optional OpenCV piece counting
    if count_pieces:
        try:
            import cv2  # noqa — just checking availability
            panels_to_count = [
                p for p in panels
                if not p.is_overall_only and p.panel_letter
                and p.pieces is None and p.photo_url
                and os.path.exists(p.photo_url)
            ]
            if panels_to_count:
                print(f"Counting pieces via OpenCV on {len(panels_to_count)} panels...")
                for pd in panels_to_count:
                    try:
                        img_bytes = open(pd.photo_url, "rb").read()
                        result = _count_pieces_opencv(img_bytes)
                        if result and result["piece_count"] > 0:
                            pd.pieces = result["piece_count"]
                    except Exception:
                        pass
        except ImportError:
            print("OpenCV not available — skipping piece counting.")

    if not panels:
        print(f"No parseable panel data found ({skipped} skipped). Creating empty sheet.")
        # Create a minimal empty sheet rather than failing
        panels = []

    print(f"Building condition sheet: {len(panels)} panels, {skipped} skipped.")
    return _build_excel(panels, output_path, project_name=project_name, flavor=flavor)
